import yfinance as yf
import pandas as pd
import numpy as np
import datetime as dt
import os
import openpyxl
import time
import math
from sklearn.linear_model import LinearRegression
from statsmodels.tsa.holtwinters import SimpleExpSmoothing
from data_functions import get_first_numeric_value, get_first_two_numeric_values

def dcf(symbol, stock_price, balance_sheet, income_stmt, cash_flow, financials_sorted, xlsx, \
        dfs, COST_OF_EQUITY, DCF_GROWTH_RATE, NUMBER_OF_YEARS, exchange_ticker, stock, filepath):
    ### Calculating Average Operating Cash Flow (ocf) Margin ###
    past_revenues = None
    ocf = None
    if os.path.exists(filepath):
        df = pd.read_excel(xlsx, sheet_name=0, index_col=0)
        past_revenues = df.loc['Revenue'][:NUMBER_OF_YEARS][::-1]
        df = pd.read_excel(xlsx, sheet_name=2, index_col=0)
        ocf = df.loc['Operating Cash Flow'][:NUMBER_OF_YEARS][::-1]

    if (past_revenues is None or past_revenues.isnull().any()) or (ocf is None or ocf.isnull().any()) or \
        (past_revenues == 0).any() or (len(past_revenues) != len(ocf)):
        try:
            past_revenues = income_stmt.loc['Total Revenue'][::-1] / 1000000
            time.sleep(0.1)
            ocf = cash_flow.loc['Operating Cash Flow'][::-1] / 1000000
            time.sleep(0.1)
        except Exception as e:
            raise Exception("Missing revenue or ocf data")


    if (past_revenues is None or past_revenues.isnull().any()) or (ocf is None or ocf.isnull().any()) or (len(past_revenues) != len(ocf)):
        raise Exception("Missing revenue or ocf data")
    else:
        try:
            past_revenues = past_revenues.astype(float)
            ocf = ocf.astype(float)
            ocf_margin = (ocf / past_revenues)
            average_ocf_margin = ocf_margin.mean()
        except Exception as e:
            raise Exception("Missing revenue or ocf data")
    
    print(f'Average ocf margin is: {average_ocf_margin}.')



    ### Calculating Average Capital Expenditure (capex) Margin ###
    capex = None
    past_revenues = None
    if os.path.exists(filepath):
        df = pd.read_excel(xlsx, sheet_name=2, index_col=0)
        capex = df.loc['Capital Expenditures'][:NUMBER_OF_YEARS][::-1]
        df = pd.read_excel(xlsx, sheet_name=0, index_col=0)
        past_revenues = df.loc['Revenue'][:NUMBER_OF_YEARS][::-1]
    
    if (past_revenues is None or past_revenues.isnull().any()) or (capex is None or capex.isnull().any()) or (len(past_revenues) != len(capex)):
        try:
            capex = cash_flow.loc['Capital Expenditure'][::-1] / 1000000
            time.sleep(0.1)    
            past_revenues = income_stmt.loc['Total Revenue'][::-1] / 1000000
            time.sleep(0.1)
        except Exception as e:
            raise Exception("Missing revenue or capex data")
        
    if (past_revenues is None or past_revenues.isnull().any()) or (capex is None or capex.isnull().any()) or (past_revenues == 0).any() or (len(past_revenues) != len(capex)):
        raise Exception("Missing revenue or capex data") 
    else:
        try:
            past_revenues = past_revenues.astype(float)
            capex = capex.astype(float)    
            capex_margin = (capex / past_revenues)
            average_capex_margin = capex_margin.mean()
        except Exception as e:
            raise Exception("Missing revenue or capex data")
        
    print(f'Average capex margin is: {average_capex_margin}.')



    ### Calculating Discount Rate (WACC/Cost of Capital) ###
    market_cap = None
    if stock_price is not None:
        try:
            shares_outstanding = None
            shares_outstanding = stock.info['sharesOutstanding'] / 1000000
            time.sleep(0.1)
            if shares_outstanding is not None or not math.isnan(shares_outstanding) or shares_outstanding != 0:
                market_cap = stock_price * shares_outstanding
            else:
                raise Exception("Missing shares outstanding data")
        except Exception as e:
            if os.path.exists(filepath):
                shares_outstanding = get_first_numeric_value(dfs, 0, 'Shares Outstanding (Diluted)')
                if shares_outstanding is None or math.isnan(shares_outstanding) or shares_outstanding == 0:
                    shares_outstanding = get_first_numeric_value(dfs, 0, 'Shares Outstanding (Basic)')
                if shares_outstanding is not None or not math.isnan(shares_outstanding) or shares_outstanding != 0:
                    market_cap = stock_price * shares_outstanding
                else:
                    market_cap = get_first_numeric_value(dfs, 11, 'Market Capitalization')
            else:
                raise Exception("Missing shares outstanding data")
    else:
        market_cap = get_first_numeric_value(dfs, 11, 'Market Capitalization')

    if market_cap is None or math.isnan(market_cap) or market_cap == 0:
        raise Exception("Missing market capitalization data")
    # else:
    #     print(f"The market capitalization is {market_cap}")

    total_debt = None
    try:
        total_debt = balance_sheet.at['Total Debt', balance_sheet.columns[0]] / 1000000
        time.sleep(0.1)
        if total_debt is None or math.isnan(total_debt) or total_debt == 0:
            raise Exception("Missing total debt data")
    except Exception as e:
        if os.path.exists(filepath):
            total_debt = get_first_numeric_value(dfs, 9, 'Total Debt')
        else:
            raise Exception("Missing total debt data")    

    if total_debt is None or math.isnan(total_debt) or total_debt == 0:
        raise Exception("Missing total debt data")
    # else:
    #     print(f"The total debt is {total_debt}")

    interest_expense = None
    try:
        interest_expense = income_stmt.at['Interest Expense', income_stmt.columns[0]] / 1000000
        time.sleep(0.1)
        if interest_expense is None or math.isnan(interest_expense) or interest_expense == 0:
            raise Exception("Missing total interest expense data")
        print(f"Data obtained from yfinance.")
        print(interest_expense)
    except Exception as e:
        if os.path.exists(filepath):
            interest_expense = get_first_numeric_value(dfs, 8, 'Interest Expense / Income')
            print(f"Data obtained from excel.")
            print(interest_expense)
        else:
            raise Exception("Missing total interest expense data")
        
    if interest_expense is None or math.isnan(interest_expense) or interest_expense == 0:
        raise Exception("Missing total interest expense data")
    # else:
    #     print(f"The total interest expense is {interest_expense}")
    
    tax_rate = None
    try:
        tax_rate = financials_sorted['Tax Rate For Calcs'].iloc[0]
        time.sleep(0.1)
        if tax_rate is None or math.isnan(tax_rate) or tax_rate == 0:
            raise Exception("Missing tax rate data")
    except Exception as e:
        try:
            tax_provision = income_stmt.at['Tax Provision', income_stmt[0]] / 1000000       # Income Tax Expense
            time.sleep(0.1)
            pretax_income = income_stmt.at['Pretax Income', income_stmt[0]] / 1000000
            time.sleep(0.1)
            if (tax_provision is not None and not math.isnan(tax_provision)) and (pretax_income is not None and \
                not math.isnan(pretax_income)) and (pretax_income != 0 and tax_provision != 0):
                tax_rate = tax_provision / pretax_income
        except Exception as e:
            if os.path.exists(filepath): 
                tax_rate = get_first_numeric_value(dfs, 0, 'Effective Tax Rate')
            else:
                raise Exception("Missing tax rate")
    
    if tax_rate is None or math.isnan(tax_rate) or tax_rate == 0:
        raise Exception("Missing tax rate")

    debt_weight = total_debt / (market_cap + total_debt)
    equity_weight = 1 - debt_weight
    cost_of_debt = interest_expense / total_debt

    wacc = debt_weight * cost_of_debt * (1 - tax_rate) + equity_weight * COST_OF_EQUITY
    print(f"The WACC is {wacc}")



    ### Forecasting Revenue using Linear Regression ###
    past_revenues = past_revenues.dropna()
    years = np.array(range(1, len(past_revenues) + 1)).reshape(-1, 1)

    model = LinearRegression()
    model.fit(years, past_revenues)

    next_years = np.array(range(len(past_revenues) + 1, len(past_revenues) + 6)).reshape(-1, 1)
    predicted_revenues = model.predict(next_years)

    print(f"The predicted revenues are:\n{predicted_revenues}")
    # for i, revenue in enumerate(predicted_revenues, start=len(past_revenues) + 1):
    #     print(f"The predicted revenue for year {i} is {revenue}")
    
    negative_revenues = False
    if any(revenue < 0 for revenue in predicted_revenues):
        model = SimpleExpSmoothing(past_revenues)
        model_fit = model.fit()
        predicted_revenues = model_fit.predict(start=len(past_revenues), end=len(past_revenues) + 4)
        negative_revenues = True

        print(f'Negative revenues, exponential smoothing predicted revenues are:\n{predicted_revenues}')

    if predicted_revenues is None or np.isnan(predicted_revenues).any():
        raise Exception("Missing predicted revenue data")

    ### Forecasting Unlevered Free Cash Flow to Firm (FCFF) ###
    projected_FCFF = []

    for revenue in predicted_revenues:
        FCFF = revenue * (average_ocf_margin - average_capex_margin)
        projected_FCFF.append(FCFF)

    print(f"The projected free cash flows are:\n{projected_FCFF}")
    # for i, FCFF in enumerate(projected_FCFF, start=len(past_revenues) + 1):
    #     print(f"The projected free cash flow for year {i} is {FCFF}")



    ### Discounting Unlevered Free Cash Flow to Firm (FCFF) ###
    # Calculate the present value of each year's free cash flow
    discounted_FCFF = [fcff / ((1 + wacc) ** year) for year, fcff in enumerate(projected_FCFF, start=1)]
    print(f"The discounted free cash flows are:\n{discounted_FCFF}")

    # for i, discounted_FCFF_i in enumerate(discounted_FCFF, start=1):
    #     print(f"The discounted free cash flow for year {i} is {discounted_FCFF_i}")



    ### Calculating the Discounted Terminal Value at the end of the projection period ###
    terminal_value = projected_FCFF[-1] * (1 + DCF_GROWTH_RATE) / (wacc - DCF_GROWTH_RATE)
    print(f"The Terminal Value is {terminal_value}")

    # Discounting Terminal Value
    forecasted_years = len(predicted_revenues)
    discounted_terminal_value = terminal_value / ((1 + wacc) ** forecasted_years)
    print(f"The Discounted Terminal Value is {discounted_terminal_value}")

    ### Calculating the Enterprise Value ###
    enterprise_value = sum(discounted_FCFF) + discounted_terminal_value
    print(f"The Enterprise Value is {enterprise_value}")



    ### Calculate Equity Value ###
    cash_and_cash_equivalents = None
    try:
        cash_and_cash_equivalents = balance_sheet.at['Cash Cash Equivalents And Short Term Investments', balance_sheet.columns[0]] / 1000000
        time.sleep(0.1)
    except Exception as e:
        if os.path.exists(filepath):
            cash_and_equivalents = get_first_numeric_value(dfs, 1, 'Cash & Equivalents')
            short_term_investments = get_first_numeric_value(dfs, 1, 'Short-TermInvestments')
            if (cash_and_equivalents is not None) and (short_term_investments is not None):
                cash_and_cash_equivalents = cash_and_equivalents + short_term_investments
            elif cash_and_equivalents is not None:
                cash_and_cash_equivalents = cash_and_equivalents
            else:
                raise Exception("Missing cash and cash equivalents data")
        else:
            raise Exception("Missing cash and cash equivalents data")
    time.sleep(0.1)

    if cash_and_cash_equivalents is None or math.isnan(cash_and_cash_equivalents):
        raise Exception("Missing cash and cash equivalents data")

    equity_value = enterprise_value + cash_and_cash_equivalents - total_debt
    print(f"The Equity Value is {equity_value}")

    ### Intrinsic Value Calculation ###
    intrinsic_value_per_share = equity_value / shares_outstanding
    print(f"The intrinsic value of {symbol} is {intrinsic_value_per_share}\n")

    workbook = openpyxl.load_workbook('valuation.xlsx')
    worksheet = workbook.active
    next_row = worksheet.max_row + 1

    worksheet.cell(row=next_row, column=1, value=exchange_ticker)
    worksheet.cell(row=next_row, column=2, value=intrinsic_value_per_share)
    if stock_price is not None:
        worksheet.cell(row=next_row, column=3, value=stock_price)
    if (stock_price is not None) and (intrinsic_value_per_share is not None):
        percentage_change = ((intrinsic_value_per_share - stock_price) / stock_price)
        cell = worksheet.cell(row=next_row, column=4, value=percentage_change)
        cell.number_format = '0.00%'
    if negative_revenues:
        worksheet.cell(row=next_row, column=5, value='DCF (Negative Revenues)')
    else:
        worksheet.cell(row=next_row, column=5, value='DCF')
    worksheet.cell(row=next_row, column=6, value=wacc)

    worksheet.cell(row=next_row, column=7, value=enterprise_value)
    worksheet.cell(row=next_row, column=8, value=equity_value)
    worksheet.cell(row=next_row, column=9, value=terminal_value)
    worksheet.cell(row=next_row, column=10, value=discounted_terminal_value)

    workbook.save('valuation.xlsx')
    with open('processed.log', 'a') as f:
        f.write(symbol + '\n')

    return intrinsic_value_per_share

    # worksheet.cell(row=next_row, column=6, value=discounted_FCFF)