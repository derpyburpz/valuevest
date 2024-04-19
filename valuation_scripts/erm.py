import yfinance as yf
import pandas as pd
import numpy as np
import datetime as dt
import os
import openpyxl
import time
import math
from sklearn.linear_model import LinearRegression
from statsmodels.tsa.holtwinters import SimpleExpSmoothing
from data_functions import get_first_numeric_value, get_first_two_numeric_values

def erm(symbol, stock_price, balance_sheet, income_stmt, xlsx, dfs, COST_OF_EQUITY, STABLE_ROE, \
        HIGH_GROWTH_PERIOD, STABLE_GROWTH_PERIOD, NUMBER_OF_YEARS, exchange_ticker, filepath, stock, STABLE_GROWTH_RATE):
    roe = None
    try:
        net_income = income_stmt.at['Net Income', income_stmt.columns[0]] / 1000000
        time.sleep(0.05)
        shareholders_equity = balance_sheet.at['Stockholders Equity', balance_sheet.columns[0]] / 1000000
        time.sleep(0.05)
        if (net_income is not None) and (shareholders_equity is not None):
            roe = net_income / shareholders_equity
        else:
            raise Exception("Missing Net Income or Shareholders Equity data")
    except Exception as e:
        if os.path.exists(filepath):
            roe = get_first_numeric_value(dfs, 11, 'Return on Equity (ROE)')
        else:
            raise Exception("Missing Net Income or Shareholders Equity data")
        
    if roe is None:
        raise Exception("Missing Net Income or Shareholders Equity data")

    total_assets = None            
    try:                       
        total_assets = balance_sheet.at['Total Assets', balance_sheet.columns[0]] / 1000000
        time.sleep(0.05)
    except Exception as e:
        if os.path.exists(filepath):
            total_assets = get_first_numeric_value(dfs, 1, 'Total Assets')
        else:
            raise Exception("Missing Total Assets data")

    if total_assets is None:
        raise Exception("Missing Total Assets data")
    
    total_liabilities = None
    try:
        total_liabilities = balance_sheet.at['Total Liabilities Net Minority Interest', balance_sheet.columns[0]] / 1000000
        time.sleep(0.05)
    except Exception as e:
        if os.path.exists(filepath):
            total_liabilities = get_first_numeric_value(dfs, 1, 'Total Liabilities')
        else:
            raise Exception("Missing Total Liabilities data")

    if total_liabilities is None:
        raise Exception("Missing Total Liabilities data")

    total_book_value_equity = total_assets - total_liabilities

    num_shares = None
    try:
        num_shares = stock.info['sharesOutstanding'] / 1000000
    except Exception as e:
        if os.path.exists(filepath):
            num_shares = get_first_numeric_value(dfs, 0, 'Shares Outstanding (Diluted)')
            if num_shares is None:
                num_shares = get_first_numeric_value(dfs, 0, 'Shares Outstanding (Basic)')
        else:
            raise Exception("Missing Shares Outstanding data")

    if num_shares is None:
        raise Exception("Missing Shares Outstanding data")
    

    # Forecasting Retained Earnings
    retained_earnings_values = None

    try:
        if os.path.exists(filepath):
            print('Using Retained Values from DFS.')
            retained_earnings_values = []
            df = pd.read_excel(xlsx, sheet_name=1, index_col=0)
            retained_earnings_values = df.loc['Retained Earnings'][:NUMBER_OF_YEARS][::-1]
        else:
            raise Exception("No excel data found for Retained Earnings or Shares Outstanding")
            
        if retained_earnings_values is None or retained_earnings_values.isnull().any() or retained_earnings_values.empty:
            raise Exception("Missing Retained Earnings or Shares Outstanding data")
            
    except Exception as e:
        try:
            retained_earnings_values = balance_sheet.loc['Retained Earnings'][::-1] / 1000000
            time.sleep(0.05)
        except Exception as e:
            raise Exception("Missing Retained Earnings or Shares Outstanding data")
            
    if retained_earnings_values is None or retained_earnings_values.isnull().any() or retained_earnings_values.empty:
        raise Exception("Missing Retained Earnings or Shares Outstanding data")

    # Forecasting Retained Earnings for the next 15 years (High Growth Period + Stable Growth Period)
    retained_earnings_values = retained_earnings_values.dropna()
    X = np.array(range(len(retained_earnings_values))).reshape(-1, 1)
    y = retained_earnings_values

    model = LinearRegression()
    model.fit(X, y)

    train_data = retained_earnings_values[:len(retained_earnings_values) - 2]
    validation_data = retained_earnings_values[-2:]

    X_train = np.array(range(len(train_data))).reshape(-1, 1)
    y_train = train_data

    model = LinearRegression()
    model.fit(X_train, y_train)

    # Generating the validation period
    X_validation = np.array(range(len(train_data), len(retained_earnings_values))).reshape(-1, 1)
    forecasted_values = model.predict(X_validation)

    forecast_errors = validation_data - forecasted_values

    # Mean Absolute Error or Root Mean Squared Error for error margin
    mae = np.mean(np.abs(forecast_errors))
    # rmse = np.sqrt(np.mean(forecast_errors**2))
    error_margin = mae

    percentage_error_margin = (error_margin / np.mean(validation_data))
    print(f'Error margin for linear regression: {percentage_error_margin}')

    # Forecast retained earnings for the next 15 years (High Growth Period + Stable Growth Period)
    forecast_period = HIGH_GROWTH_PERIOD + STABLE_GROWTH_PERIOD
    X_forecast = np.array(range(len(retained_earnings_values), len(retained_earnings_values) + forecast_period)).reshape(-1, 1)
    forecasted_retained_earnings = model.predict(X_forecast)


    # Stage 1: High Growth Period
    book_value_equity_per_share = total_book_value_equity / num_shares
    excess_returns_per_share = []
    for year in range(1, HIGH_GROWTH_PERIOD + 1):
        # Calculate excess return per share
        excess_return_per_share = book_value_equity_per_share * (roe - COST_OF_EQUITY) * (1 + STABLE_GROWTH_RATE)
        excess_returns_per_share.append(excess_return_per_share)

        # Update book value of equity per share for the next year
        book_value_equity_per_share += forecasted_retained_earnings[year - 1] / num_shares

        # Discount the excess returns
        discounted_excess_returns = [excess_return / ((1 + COST_OF_EQUITY) ** year) for year, excess_return in enumerate(excess_returns_per_share, start=1)]

    # Stage 2: Stable Growth Period
    terminal_year_excess_return = book_value_equity_per_share * (STABLE_ROE - COST_OF_EQUITY)
    excess_returns_terminal_stage = terminal_year_excess_return * (COST_OF_EQUITY - STABLE_GROWTH_RATE)
    discounted_excess_return_terminal_stage = excess_returns_terminal_stage / ((1 + COST_OF_EQUITY) ** HIGH_GROWTH_PERIOD)


    # Calculating the Estimated Value
    estimated_value = sum(discounted_excess_returns) + discounted_excess_return_terminal_stage + book_value_equity_per_share
    lower_bound = estimated_value - error_margin
    upper_bound = estimated_value + error_margin
    print(f"The estimated value of {symbol} is {estimated_value} (Range: {lower_bound} to {upper_bound})\n")

    # Write results to Excel
    workbook = openpyxl.load_workbook('valuation.xlsx')
    worksheet = workbook.active
    next_row = worksheet.max_row + 1

    percentage_change = (estimated_value - stock_price) / stock_price
    cell = worksheet.cell(row=next_row, column=4, value=percentage_change)
    cell.number_format = '0.00%'

    worksheet.cell(row=next_row, column=1, value=exchange_ticker)
    worksheet.cell(row=next_row, column=2, value=percentage_error_margin)
    cell = worksheet.cell(row=next_row, column=2)
    cell.number_format = '0.00%'
    # worksheet.cell(row=next_row, column=2, value=error_margin)
    worksheet.cell(row=next_row, column=3, value=estimated_value)
    if stock_price is not None:
        worksheet.cell(row=next_row, column=4, value=stock_price)
    if (stock_price is not None) and (estimated_value is not None):
        percentage_change = ((estimated_value - stock_price) / stock_price)
        cell = worksheet.cell(row=next_row, column=5, value=percentage_change)
        cell.number_format = '0.00%'
    if HIGH_GROWTH_PERIOD > 1:
        print('HIGH GROWTH WRITTEN TO EXCEL.')
        worksheet.cell(row=next_row, column=6, value='ERM (High-Growth)')
    else:
        print('MATURE WRITTEN TO EXCEL.')
        worksheet.cell(row=next_row, column=6, value='ERM (Mature)')
    worksheet.cell(row=next_row, column=14, value=excess_returns_terminal_stage)
    worksheet.cell(row=next_row, column=15, value=discounted_excess_return_terminal_stage)
    discountedExcessReturnPerShare_str = ", ".join(map(str, discounted_excess_returns.tolist()))
    worksheet.cell(row=next_row, column=16, value=discountedExcessReturnPerShare_str)

    workbook.save('valuation.xlsx')
    with open('processed.log', 'a') as f:
        f.write(symbol + '\n')

    return estimated_value